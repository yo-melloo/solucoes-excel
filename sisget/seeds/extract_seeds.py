"""
Extrai dados de seed da planilha SISGET para JSON.
Arquivos gerados:
  - motoristas.json  (BANCO DE DADOS: col A-C)
  - frota.json       (BANCO DE DADOS: col E-F)
  - ravs.json        (CONTROLE DE RAVs: col B-H)
  - reservas.json    (RESERVAS E QUARTOS: col B-E, oficina)
"""
import json
import openpyxl
import os
from pathlib import Path

# Configuração de Caminhos Dinâmicos
# Script está em: workspace/sisget/seeds/extract_seeds.py
BASE_DIR = Path(__file__).resolve().parent  # sisget/seeds/
PROJECT_ROOT = BASE_DIR.parent.parent       # workspace/

# Planilha de Origem na pasta vizinha
FILE = PROJECT_ROOT / "Satélite Norte" / "SISGET - SISTEMA DE GERENCIAMENTO DE TRÁFEGO.xlsx"
# Saída na própria pasta do script
OUT_DIR = BASE_DIR

if not FILE.exists():
    print(f"[ERRO] Planilha não encontrada em: {FILE}")
    exit(1)

print(f"[*] Abrindo planilha: {FILE.name}")
wb = openpyxl.load_workbook(FILE, data_only=True)

# ---- MOTORISTAS + FROTA (aba "BANCO DE DADOS") ----
print("[*] Extraindo Motoristas e Frota...")
ws = wb['BANCO DE DADOS']
motoristas = []
frota = []
seen_mat = set()
seen_frota = set()

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    nome = row[0].value  # A - MOTORISTA
    mat  = row[1].value  # B - MATRICULA
    if nome and mat and mat not in seen_mat:
        motoristas.append({
            "nome": str(nome).strip(),
            "matricula": int(mat) if isinstance(mat, (int, float)) else str(mat).strip(),
        })
        seen_mat.add(mat)

    veiculo = row[4].value  # E - Veículo (prefixo)
    placa   = row[5].value  # F - Placa
    if veiculo and placa and veiculo not in seen_frota:
        frota.append({
            "prefixo": str(veiculo).strip(),
            "placa": str(placa).strip(),
        })
        seen_frota.add(veiculo)

# ---- RAVs (aba "CONTROLE DE RAVs") ----
print("[*] Extraindo RAVs...")
ws_rav = wb['CONTROLE DE RAVs']
ravs = []
for row in ws_rav.iter_rows(min_row=3, max_row=ws_rav.max_row, values_only=False):
    chamado = row[2].value  # C
    carro   = row[3].value  # D
    envolvido = row[4].value  # E
    escopo  = row[5].value  # F
    doc     = row[6].value  # G
    status  = row[7].value  # H
    if chamado:
        ravs.append({
            "chamado": str(chamado).strip(),
            "carro": str(carro).strip() if carro else None,
            "envolvido": str(envolvido).strip() if envolvido else None,
            "escopo": str(escopo).strip() if escopo else None,
            "documentacao": str(doc).strip() if doc else None,
            "status": str(status).strip() if status else None,
        })

# ---- RESERVAS (aba "RESERVAS E QUARTOS", seção Oficina) ----
print("[*] Extraindo Reservas...")
ws_res = wb['RESERVAS E QUARTOS']
reservas = []
for row in ws_res.iter_rows(min_row=5, max_row=ws_res.max_row, values_only=False):
    cod  = row[1].value  # B - CÓD
    tipo = row[2].value  # C - TIPO
    stat = row[3].value  # D - STATUS
    desc = row[4].value  # E - DESCRIÇÃO
    if cod and tipo:
        reservas.append({
            "codigo": str(cod).strip(),
            "tipo": str(tipo).strip(),
            "status": str(stat).strip() if stat else None,
            "descricao": str(desc).strip() if desc else None,
        })

# ---- WRITE ----
def write_json(data, name):
    target = OUT_DIR / f"{name}.json"
    with open(target, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    print(f"  [OK] {name}.json: {len(data)} registros")

print("\nGerando arquivos JSON:")
write_json(motoristas, "motoristas")
write_json(frota, "frota")
write_json(ravs, "ravs")
write_json(reservas, "reservas")
print("\nConcluído com sucesso!")
